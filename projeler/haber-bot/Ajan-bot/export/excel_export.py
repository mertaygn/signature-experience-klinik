"""
Excel/CSV Export
Export enriched company data to Excel and CSV formats.
"""

import csv
from datetime import datetime
from pathlib import Path
from typing import Optional
from rich.console import Console

import config
from database.db import Database

console = Console()


def export_to_csv(db: Database, fair_id: int, output_path: Optional[Path] = None) -> Path:
    """
    Export enriched company data to CSV.

    Returns path to the created file.
    """
    fair = None
    fairs = db.get_all_fairs()
    for f in fairs:
        if f["id"] == fair_id:
            fair = f
            break

    fair_name = fair["slug"] if fair else f"fair_{fair_id}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if not output_path:
        output_path = config.EXPORT_DIR / f"{fair_name}_{timestamp}.csv"

    companies = db.get_all_enriched_data(fair_id)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            "Firma Adı", "Web Sitesi", "Stand No", "Sektör", "Ülke", "Şehir",
            "E-postalar", "Telefonlar", "Kişiler", "LinkedIn", "Twitter",
            "Instagram", "Adres", "Açıklama"
        ])

        for company in companies:
            contacts = company.get("contacts", [])

            # Group contacts by type
            emails = [c["value"] for c in contacts if c["contact_type"] == "email"]
            phones = [c["value"] for c in contacts if c["contact_type"] == "phone"]
            people = [f"{c['value']} ({c.get('label', '')})" for c in contacts if c["contact_type"] == "person"]
            social = {c["label"]: c["value"] for c in contacts if c["contact_type"] == "social"}
            addresses = [c["value"] for c in contacts if c["contact_type"] == "address"]

            writer.writerow([
                company["name"],
                company.get("website", ""),
                company.get("booth_number", ""),
                company.get("sector", ""),
                company.get("country", ""),
                company.get("city", ""),
                " | ".join(emails),
                " | ".join(phones),
                " | ".join(people),
                social.get("linkedin", ""),
                social.get("twitter", ""),
                social.get("instagram", ""),
                " | ".join(addresses),
                company.get("description", ""),
            ])

    console.print(f"[green]✓ CSV exported: {output_path}[/green]")
    console.print(f"  [dim]{len(companies)} companies exported[/dim]")
    return output_path


def export_to_excel(db: Database, fair_id: int, output_path: Optional[Path] = None) -> Path:
    """
    Export enriched company data to Excel with formatting.

    Returns path to the created file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        console.print("[yellow]openpyxl not installed. Falling back to CSV.[/yellow]")
        return export_to_csv(db, fair_id, output_path)

    fair = None
    fairs = db.get_all_fairs()
    for f in fairs:
        if f["id"] == fair_id:
            fair = f
            break

    fair_name = fair["slug"] if fair else f"fair_{fair_id}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if not output_path:
        output_path = config.EXPORT_DIR / f"{fair_name}_{timestamp}.xlsx"

    companies = db.get_all_enriched_data(fair_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Firma Listesi"

    # ─── Styles ───────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    alt_fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")

    # ─── Headers ──────────────────────────────────────────────
    headers = [
        "Firma Adı", "Web Sitesi", "Stand No", "Sektör", "Ülke",
        "E-postalar", "Telefonlar", "Kişiler & Pozisyonlar",
        "LinkedIn", "Adres"
    ]
    col_widths = [30, 30, 12, 20, 15, 35, 25, 35, 35, 40]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ─── Data Rows ────────────────────────────────────────────
    for row_idx, company in enumerate(companies, 2):
        contacts = company.get("contacts", [])

        emails = [c["value"] for c in contacts if c["contact_type"] == "email"]
        phones = [c["value"] for c in contacts if c["contact_type"] == "phone"]
        people = [f"{c['value']} — {c.get('label', '')}" for c in contacts if c["contact_type"] == "person"]
        social = {c["label"]: c["value"] for c in contacts if c["contact_type"] == "social"}
        addresses = [c["value"] for c in contacts if c["contact_type"] == "address"]

        row_data = [
            company["name"],
            company.get("website", ""),
            company.get("booth_number", ""),
            company.get("sector", ""),
            company.get("country", ""),
            "\n".join(emails),
            "\n".join(phones),
            "\n".join(people),
            social.get("linkedin", ""),
            "\n".join(addresses),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value or "")
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

            # Make website and LinkedIn clickable
            if col in [2, 9] and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    # ─── Summary Sheet ────────────────────────────────────────
    ws2 = wb.create_sheet("Özet")
    stats = db.get_stats(fair_id)

    summary_data = [
        ("Fuar", fair["name"] if fair else ""),
        ("Toplam Firma", stats["total_companies"]),
        ("E-postası Bulunan", stats["with_email"]),
        ("Telefonu Bulunan", stats["with_phone"]),
        ("Zenginleştirilen", stats["enriched"]),
        ("Toplam Kontak", stats["total_contacts"]),
        ("Export Tarihi", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    for row, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=value)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 40

    # ─── Freeze & Filter ─────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    console.print(f"[green]✓ Excel exported: {output_path}[/green]")
    console.print(f"  [dim]{len(companies)} companies with {stats['total_contacts']} contacts[/dim]")
    return output_path
